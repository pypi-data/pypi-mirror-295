
import os
import traceback
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rogue_tools import path_tool,json_tool,file_tool

class RogueSheet():
    '''
    仅保留有效数据组成的矩阵,获取行或列的时候,请从第一行数据开始,从0开始
    处理表头,用专门的方法
    '''
    def __init__(self,sheet_name,sheet:dict) -> None:
        self.name               = sheet_name
        self.key_list           = sheet['key_list']
        self.max_row            = sheet['max_row']
        self.rows               = sheet['rows']
        self.columns             = {key:[] for key in self.key_list}
        for row in self.rows:
            for index in range(len(row)):
                self.columns[self.key_list[index]].append(row[index])

    def get_keys(self):
        return self.key_list

    def get_value(self,row_index,column_index):
        return self.rows[row_index][column_index]

    def get_rows(self):
        '''取得所有row'''
        return self.rows

    def get_row_list(self,index):
        '''取得某一行'''
        if index >= self.max_row or index < 0:
            return []
        return self.rows[index]

    def get_column_list(self,key):
        '''取得某一列'''
        if type(key) == int:
            key = self.key_list[key]
        return self.columns.get(key,None)
        
class RogueExcel():
    def __init__(self,excel_file_path,using_cache=True) -> None:
        self.excel_file_path                            = excel_file_path
        self.excel_file_cache_path                      = f'{excel_file_path}.json'
        self._work_book                                 = None
        self.sheet_jsonobj_dic                          = {}
        self.sheet_dic:dict[str,RogueSheet]             = {}
        self._sheet_name_list                           = []
        self.using_cache                                = using_cache
        self.init_by_file()
        self.change_json_2_sheet()
        print('加载完成')
        
    def init_by_file(self):
        '''
        读取文件转为json格式
        '''
        if os.path.exists(self.excel_file_cache_path) and self.using_cache:
            self.load_by_json()
        elif os.path.exists(self.excel_file_path):
            self.load_by_excel()
        else:
            raise FileNotFoundError('不存在',self.excel_file_path,self.excel_file_cache_path)

    def change_json_2_sheet(self):
        '''将json格式转为sheet'''
        for sheet_name , sheet_json in self.sheet_jsonobj_dic.items():
            if sheet_name != 'last_excel_md5':
                self.sheet_dic[sheet_name]=RogueSheet(sheet_name,sheet_json)

    def load_by_json(self):
        if not os.path.exists(self.excel_file_cache_path) or not self.using_cache:
            self.load_by_excel()
            return
        md5 = file_tool.get_md5(self.excel_file_path)
        print(f'加载Json:{self.excel_file_path}')
        self.sheet_jsonobj_dic = json_tool.load_json_by_file(self.excel_file_cache_path)
        last_md5 = self.sheet_jsonobj_dic['last_excel_md5']
        # 文件发生变化,更新json
        if md5!=last_md5:
            path_tool.del_(self.excel_file_cache_path)
            self.load_by_excel()
            
    def load_by_excel(self):
        if not os.path.exists(self.excel_file_path):
            return
        print(f'加载Excel:{self.excel_file_path}')
        try:
            self._work_book:openpyxl.Workbook   = openpyxl.load_workbook(self.excel_file_path) # 打开Excel文件
            self._sheet_name_list               = self._work_book.sheetnames
            for sheet_name in self._sheet_name_list:
                sheet:Worksheet = self._work_book[sheet_name]
                rows = []
                # 加载数据
                for row in sheet.rows:
                    row_value = []
                    for cell in row:
                        value = cell.value if cell.value or cell.value == 0 else ''
                        row_value.append(value)
                    rows.append(row_value)
                self.sheet_jsonobj_dic[sheet_name] = self.__new_sheet_json_obj(sheet_name,rows)
        except Exception:
            traceback.print_exc()
            print(f'加载Excel文件失败:{self.excel_file_path}')
        finally:
            self._work_book.close()
            self.sheet_jsonobj_dic['last_excel_md5'] = file_tool.get_md5(self.excel_file_path)
            json_tool.save_json_by_dict(self.excel_file_cache_path,self.sheet_jsonobj_dic)
            print(f'已关闭文件:{self.excel_file_path}')


    def get_sheet_name_list(self):
        return self._sheet_name_list

    def __new_sheet_json_obj(self,sheet_name,rows):
        sheet_obj = {}
        if rows:
            sheet_obj['key_list'] = rows[0]
            sheet_obj['max_row'] = len(rows[0])
            sheet_obj['rows'] = rows
        else:
            print('发现空列表:',sheet_name)
            sheet_obj['key_list'] = []
            sheet_obj['max_row'] = 0
            sheet_obj['rows'] = []
        return sheet_obj

    def get_sheet(self,sheet_name):
        return self.sheet_dic.get(sheet_name,None)
    
    def get_key_list(self,sheet_name):
        return self.get_sheet(sheet_name)['key_list']

    def get_max_row(self,sheet_name):
        return self.get_sheet(sheet_name)['max_row']

    def get_rows(self,sheet_name):
        return self.get_sheet(sheet_name)['rows']

    def get_row(self,sheet_name,row_index):
        return self.get_rows(sheet_name)[row_index]
