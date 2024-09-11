
import os
import traceback
import math

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from rogue_tools import path_tool,json_tool,file_tool



def standard_list(src_list,list_len,standard_obj=None):
    '''
    长的截断,短的填充standard_obj
    '''
    if len(src_list)<list_len:
        for i in range(0,list_len-len(src_list)):
            src_list.append(standard_obj)
    
    return src_list[0:list_len]


def is_in_filter(src_str,include_list=[],exclude_list=[]):
    '''
    检查src_str是否符合筛选条件(str类型)
    '''
    if include_list==[] and exclude_list==[]:
        return True

    for find_str in exclude_list:
        if src_str.find(find_str)>-1:
            return False
    for find_str in include_list:
        if src_str.find(find_str)>-1:
            return True
    return False

def calc_26_to_10(a_z:str):
    rs=0
    a_z = a_z.lower()
    sq = len(a_z)-1
    for word in a_z:
        number = (ord(word)-ord('a')+1)*math.pow(26,sq)
        sq=sq-1
        rs = rs+number
    return int(rs)



def cut_list_tail(src_list,list_obj):
    for i in range(0,len(src_list)):
        index = len(src_list)-i-1
        if src_list[index]!=list_obj:
            return src_list[0:index+1]
    return []


class MySheet():
    '''
    仅保留有效数据组成的矩阵,获取行或列的时候,请从第一行数据开始,从0开始
    处理表头,用专门的方法
    '''
    def __init__(self,sheet:Worksheet) -> None:
        if sheet == None:
            return
        #self.start_line         = 0
        self.error_list         = []
        self.row_list           = [] # 是一个规整的二维矩阵
        self.column_list        = [] # 是一个规整的二维矩阵
        self.key_list           = []
        self.name               = sheet.title
        if sheet.max_row==0 or sheet.max_column==0:
            return
        
        # 加载数据
        for row in sheet.rows:
            row_value = []
            for cell in row:
                value = cell.value if cell.value or cell.value == 0 else ''
                row_value.append(value)
            self.row_list.append(row_value)
        
        for column in sheet.columns:
            column_value = []
            for cell in column:
                value = cell.value if cell.value or cell.value == 0 else ''
                column_value.append(value)
            self.column_list.append(column_value)
        # 标记序列的行
        self.key_list = cut_list_tail(self.row_list[0],'')
        self.index_list = cut_list_tail(self.column_list[0],'')
        # 获得最大行列
        self.max_row = len(self.index_list)
        self.max_column= len(self.key_list)
        # 重新规整数据
        del self.row_list[self.max_row:]
        for i in range(0,len(self.row_list)-1):
            del self.row_list[i][self.max_column:]
        
        #print(f'init {self.name}')

    def get_value(self,x,y):
        '''
        通过表内坐标获得value,从0,0开始
        '''
        return self.row_list[x][y]
    def get_all_rows(self):
        return self.row_list
    def get_row_list(self,key):
        '''
        取得某一行，从0开始
        '''
        if key >= self.max_row or key < 0:
            return None
        elif key in self.column_list[0]:
            return self.row_list[self.column_list[0].index(key)]
        else:
            return self.row_list[key]

    def get_column_list(self,key):
        '''
        取得某一列
        '''

        if type(key)==int:
            if key >= self.max_column or key < 0:
                return None
            else:
                return self.column_list[key]
        elif key in self.row_list[0]:
            return self.column_list[self.row_list[0].index(key)]
        else:
            return self.get_column_list(calc_26_to_10(key)-1)

    def change_list(self):
        '''将表转换为字典的形式存储'''


class MyExcel():
    def __init__(self,excel_file_path) -> None:
        self.excel_file_path             = excel_file_path
        self._work_book                  = None
        self.sheet_dic:dict[str,MySheet]     = {}
        self._sheet_name_list            = []
        self.is_read                     = False
        self.read_file()

    def read_file(self):
        if os.path.exists(self.excel_file_path):
            try:
                self._work_book:openpyxl.Workbook   = openpyxl.load_workbook(self.excel_file_path) # 打开Excel文件
                self._sheet_name_list               = self._work_book.sheetnames
                for sheet_name in self._sheet_name_list:
                    self.sheet_dic[sheet_name] = MySheet(self._work_book[sheet_name])
                self.is_read = True
            except Exception:
                traceback.print_exc()
                print(f'加载Excel文件失败:{self.excel_file_path}')
            finally:
                self._work_book.close()
                print(f'已关闭文件:{self.excel_file_path}')
        else:
            print('不存在',self.excel_file_path)
    def write_excel(self,other_file=None,is_backup=True):
        '''
        修改这个Myexcel
        '''
        if os.path.exists(self.excel_file_path):
            self._work_book = openpyxl.load_workbook(self.excel_file_path)
        else:
            self._work_book = openpyxl.Workbook()

        try:
            for sheet_name in self.sheet_dic:
                sheet_row_value = self.sheet_dic[sheet_name].row_list # 按行分配的单元格列表
                if not sheet_row_value:
                    sheet_row_value=[['']]

                sheet = self._work_book[sheet_name]
                for row_index in range(0,len(sheet_row_value)):
                    row_value = sheet_row_value[row_index]
                    for col_index in range(0,len(row_value)):
                        sheet.cell(row_index+1,col_index+1,f'{row_value[col_index]}')

            save_path = self.excel_file_path
            if other_file:
                save_path = other_file
            if is_backup:
                path_tool.copy(self.excel_file_path,self.excel_file_path+'.bak')
            self._work_book.save(save_path)
            
        except Exception:
            traceback.print_exc()
        finally:
            self._work_book.close()
            
            

    
    def get_sheet_name_list(self):
        return self._sheet_name_list

    def get_sheet(self,sheet_name) -> MySheet:
        return self.sheet_dic.get(sheet_name,None)
    
    def add_sheet(self,sheet_name)-> MySheet:
        self._sheet_name_list.append(sheet_name)
        self.sheet_dic[sheet_name]=MySheet(sheet_name)
        return self.sheet_dic[sheet_name]

    def save_cache(self):
        '''将这个excel转为json格式的缓存文件，方便之后读取'''
        save_dic = {}
        for sheet_name in self.sheet_dic:
            sheet = self.sheet_dic[sheet_name]
            sheet.get_all_rows()
            save_dic[sheet_name] = sheet.get_all_rows()
        json_str = json_tool.output_json_by_dict(save_dic)
        file_tool.write_str(f'{self.excel_file_path}.json',json_str,'w+')


def calc_26_to_10(a_z:str):
    rs=0
    a_z = a_z.lower()
    sq = len(a_z)-1
    for word in a_z:
        number = (ord(word)-ord('a')+1)*math.pow(26,sq)
        sq=sq-1
        rs = rs+number
    return int(rs)

